from django.shortcuts import render
import pandas as pd
from datetime import datetime

import os
import qrcode
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import load_workbook

# Create your views here.
def dashboard(request):
    table_of_summary = dashboard_summary_of_assets(request)
    yearly_table=yearly_summary(request)
    context = {'table': table_of_summary,
               'yearly': yearly_table}
    
    
    excel_file_path_qr_generation = 'csv_path/excel_files/asset_register.xlsx'

    # Delete previously generated QR codes
    qr_code_directory = os.path.join(settings.QR_MEDIA_ROOT, 'qrcode_saved_folder')
    if os.path.exists(qr_code_directory):
        for file in os.listdir(qr_code_directory):
            os.remove(os.path.join(qr_code_directory, file))
    else:
        os.makedirs(qr_code_directory)

    # Generate new QR codes from Excel data
    wb = load_workbook(excel_file_path_qr_generation)
    ws = wb.active
    qr_codes = []
    skip_first_row = True
    for row in ws.iter_rows(values_only=True):
        if skip_first_row:
            skip_first_row = False
            continue
        asset_code = row[4]  # Assuming asset code is in the first column
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(asset_code)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img_path = os.path.join(qr_code_directory, f"{asset_code}.png")
        img.save(img_path)
        qr_codes.append(img_path)

    return render(request, "Dashboard.html", context)

def dashboard_summary_of_assets(request):
    file_path ='csv_path/excel_files/asset_register.xlsx'
     # Replace with the actual file path
    df = pd.read_excel(file_path)
    list_primary=df[['Category','Expected life','Modified Number']]
    list_primary=list_primary['Category'].drop_duplicates().dropna()
    list_count=list_primary
    total_bills=0
    total_items=0
    table=pd.DataFrame(columns=['Category','Number of Instances (Bills)','Number of Items'])
    print(table)
    for i in list_count:
        j= df[df['Category'] == i].shape[0]
        t= df[df['Category'] == i]['Modified Number'].sum()
        new_row = pd.DataFrame({'Category': [i], 'Number of Instances (Bills)': [j],'Number of Items':[t]})
        table = pd.concat([table, new_row], ignore_index=True)
        total_bills=total_bills+j
        total_items=total_items+t
    new_list=pd.merge(table,list_primary,on='Category',how='left')
    new_row = pd.DataFrame({'Category': "TOTAL " , 'Number of Instances (Bills)': [total_bills],"Number of Items":[total_items]})
    table = pd.concat([new_list, new_row], ignore_index=True)
    table=table.fillna(' ')
    table_of_summary=table.to_html(index=False)
    return table_of_summary


def yearly_summary(request):
    file_path ='csv_path/excel_files//asset_register.xlsx'
    df = pd.read_excel(file_path)
    list_primary=df['Financial Year']
    list_primary=list_primary.drop_duplicates().dropna()
    list_count=list_primary
    total_bills=0
    total_items=0
    table=pd.DataFrame(columns=['Financial Year','Number of Instances (Bills)','Number of Items'])
    for i in list_count:
        j= df[df['Financial Year'] == i].shape[0]
        t= df[df['Financial Year'] == i]['Modified Number'].sum()
        new_row = pd.DataFrame({'Financial Year': [i], 'Number of Instances (Bills)': [j],'Number of Items':[t]})
        table = pd.concat([table, new_row], ignore_index=True)
        total_bills=total_bills+j
        total_items=total_items+t
    new_list=table
    new_row = pd.DataFrame({'Financial Year': "TOTAL " , 'Number of Instances (Bills)': [total_bills],"Number of Items":[total_items]})
    table = pd.concat([new_list, new_row], ignore_index=True)
    table=table.fillna(' ')
    table=table.to_html(index=(False))
    return table